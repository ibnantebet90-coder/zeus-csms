"""
ZEUS CSMS — Report Endpoints
GET  /api/reports/daily          — laporan harian (per CP, per hari)
GET  /api/reports/monthly        — laporan bulanan (agregasi per bulan)
GET  /api/reports/summary        — ringkasan untuk periode tertentu
GET  /api/reports/export/csv     — export laporan ke CSV
GET  /api/reports/export/excel   — export laporan ke Excel (.xlsx)
"""

from datetime import datetime, date, timedelta
from typing import List, Optional
from io import BytesIO

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, extract
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.core.database import get_db
from app.models.models import ChargePoint, Transaction, User
from app.schemas.schemas import ReportDaily, ReportMonthly, ReportSummary

router = APIRouter(tags=["Reports"])


# ── Helpers ───────────────────────────────────────────────────────────────────

def _base_query(db: Session, date_from: date, date_to: date, charge_point_id: Optional[str]):
    """Query transaksi Completed dalam rentang tanggal."""
    q = db.query(Transaction).filter(
        Transaction.status == "Completed",
        Transaction.start_timestamp >= datetime.combine(date_from, datetime.min.time()),
        Transaction.start_timestamp <= datetime.combine(date_to, datetime.max.time()),
    )
    if charge_point_id:
        q = q.filter(Transaction.charge_point_id == charge_point_id)
    return q


# ── Laporan Harian ────────────────────────────────────────────────────────────

@router.get("/api/reports/daily", response_model=List[ReportDaily])
def report_daily(
    date_from: date     = Query(default_factory=lambda: date.today() - timedelta(days=6)),
    date_to:   date     = Query(default_factory=date.today),
    charge_point_id: Optional[str] = Query(None),
    db: Session         = Depends(get_db),
    _: User             = Depends(get_current_user),
):
    """
    Agregasi per hari + per charge point.
    Kembalikan list: tanggal, cp_id, jumlah tx, energi (kWh), pendapatan (IDR).
    """
    rows = (
        _base_query(db, date_from, date_to, charge_point_id)
        .with_entities(
            func.date(Transaction.start_timestamp).label("report_date"),
            Transaction.charge_point_id,
            func.count(Transaction.id).label("total_transactions"),
            func.coalesce(func.sum(Transaction.energy_consumed_kwh), 0).label("total_energy_kwh"),
            func.coalesce(func.sum(Transaction.total_cost), 0).label("total_revenue"),
        )
        .group_by(
            func.date(Transaction.start_timestamp),
            Transaction.charge_point_id,
        )
        .order_by(func.date(Transaction.start_timestamp).desc())
        .all()
    )

    # Ambil nama CP sekali saja
    cp_names: dict[str, str] = {
        cp.charge_point_id: cp.name
        for cp in db.query(ChargePoint.charge_point_id, ChargePoint.name).all()
    }

    return [
        ReportDaily(
            report_date=str(r.report_date),
            charge_point_id=r.charge_point_id,
            charge_point_name=cp_names.get(r.charge_point_id, r.charge_point_id),
            total_transactions=r.total_transactions,
            total_energy_kwh=float(r.total_energy_kwh),
            total_revenue=float(r.total_revenue),
        )
        for r in rows
    ]


# ── Laporan Bulanan ───────────────────────────────────────────────────────────

@router.get("/api/reports/monthly", response_model=List[ReportMonthly])
def report_monthly(
    year: int                      = Query(default_factory=lambda: datetime.today().year),
    charge_point_id: Optional[str] = Query(None),
    db: Session                    = Depends(get_db),
    _: User                        = Depends(get_current_user),
):
    """
    Agregasi per bulan dalam satu tahun.
    """
    date_from = date(year, 1, 1)
    date_to   = date(year, 12, 31)

    rows = (
        _base_query(db, date_from, date_to, charge_point_id)
        .with_entities(
            extract("year",  Transaction.start_timestamp).label("year"),
            extract("month", Transaction.start_timestamp).label("month"),
            Transaction.charge_point_id,
            func.count(Transaction.id).label("total_transactions"),
            func.coalesce(func.sum(Transaction.energy_consumed_kwh), 0).label("total_energy_kwh"),
            func.coalesce(func.sum(Transaction.total_cost), 0).label("total_revenue"),
            func.coalesce(func.avg(Transaction.energy_consumed_kwh), 0).label("avg_energy_kwh"),
        )
        .group_by(
            extract("year",  Transaction.start_timestamp),
            extract("month", Transaction.start_timestamp),
            Transaction.charge_point_id,
        )
        .order_by(
            extract("year",  Transaction.start_timestamp).desc(),
            extract("month", Transaction.start_timestamp).desc(),
        )
        .all()
    )

    cp_names: dict[str, str] = {
        cp.charge_point_id: cp.name
        for cp in db.query(ChargePoint.charge_point_id, ChargePoint.name).all()
    }

    month_names = [
        "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
        "Juli", "Agustus", "September", "Oktober", "November", "Desember",
    ]

    return [
        ReportMonthly(
            year=int(r.year),
            month=int(r.month),
            month_name=month_names[int(r.month)],
            charge_point_id=r.charge_point_id,
            charge_point_name=cp_names.get(r.charge_point_id, r.charge_point_id),
            total_transactions=r.total_transactions,
            total_energy_kwh=float(r.total_energy_kwh),
            total_revenue=float(r.total_revenue),
            avg_energy_kwh=float(r.avg_energy_kwh),
        )
        for r in rows
    ]


# ── Ringkasan Periode ─────────────────────────────────────────────────────────

@router.get("/api/reports/summary", response_model=ReportSummary)
def report_summary(
    date_from: date     = Query(default_factory=lambda: date.today() - timedelta(days=29)),
    date_to:   date     = Query(default_factory=date.today),
    charge_point_id: Optional[str] = Query(None),
    db: Session         = Depends(get_db),
    _: User             = Depends(get_current_user),
):
    """
    Agregasi total untuk summary card di atas halaman report.
    """
    q = _base_query(db, date_from, date_to, charge_point_id)

    total_tx, total_energy, total_revenue = q.with_entities(
        func.count(Transaction.id),
        func.coalesce(func.sum(Transaction.energy_consumed_kwh), 0),
        func.coalesce(func.sum(Transaction.total_cost), 0),
    ).one()

    avg_energy, avg_cost = q.with_entities(
        func.coalesce(func.avg(Transaction.energy_consumed_kwh), 0),
        func.coalesce(func.avg(Transaction.total_cost), 0),
    ).one()

    active_cp = (
        q.with_entities(func.count(func.distinct(Transaction.charge_point_id)))
        .scalar() or 0
    )

    return ReportSummary(
        date_from=str(date_from),
        date_to=str(date_to),
        total_transactions=total_tx or 0,
        total_energy_kwh=float(total_energy),
        total_revenue=float(total_revenue),
        avg_energy_per_tx_kwh=float(avg_energy),
        avg_cost_per_tx=float(avg_cost),
        active_charge_points=active_cp,
    )


# ── Export CSV ────────────────────────────────────────────────────────────────

@router.get("/api/reports/export/csv")
def export_report_csv(
    report_type: str    = Query("daily", regex="^(daily|monthly)$"),
    date_from: date     = Query(default_factory=lambda: date.today() - timedelta(days=29)),
    date_to:   date     = Query(default_factory=date.today),
    year: int           = Query(default_factory=lambda: datetime.today().year),
    charge_point_id: Optional[str] = Query(None),
    db: Session         = Depends(get_db),
    _: User             = Depends(get_current_user),
):
    import csv, io

    output = io.StringIO()
    writer = csv.writer(output)

    if report_type == "daily":
        writer.writerow(["Tanggal", "Charge Point ID", "Nama CP",
                         "Jumlah Transaksi", "Energi (kWh)", "Pendapatan (Rp)"])
        rows = report_daily(date_from, date_to, charge_point_id, db, _)
        for r in rows:
            writer.writerow([
                r.report_date, r.charge_point_id, r.charge_point_name,
                r.total_transactions,
                f"{r.total_energy_kwh:.3f}",
                f"{r.total_revenue:.0f}",
            ])
        filename = f"laporan_harian_{date_from}_{date_to}.csv"
    else:
        writer.writerow(["Tahun", "Bulan", "Charge Point ID", "Nama CP",
                         "Jumlah Transaksi", "Energi (kWh)", "Rata-rata kWh/tx", "Pendapatan (Rp)"])
        rows = report_monthly(year, charge_point_id, db, _)
        for r in rows:
            writer.writerow([
                r.year, r.month_name, r.charge_point_id, r.charge_point_name,
                r.total_transactions,
                f"{r.total_energy_kwh:.3f}",
                f"{r.avg_energy_kwh:.3f}",
                f"{r.total_revenue:.0f}",
            ])
        filename = f"laporan_bulanan_{year}.csv"

    output.seek(0)
    # Tambahkan BOM agar Excel bisa baca UTF-8
    content = "\uFEFF" + output.getvalue()
    return StreamingResponse(
        iter([content.encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Export Excel ──────────────────────────────────────────────────────────────

@router.get("/api/reports/export/excel")
def export_report_excel(
    report_type: str    = Query("daily", regex="^(daily|monthly)$"),
    date_from: date     = Query(default_factory=lambda: date.today() - timedelta(days=29)),
    date_to:   date     = Query(default_factory=date.today),
    year: int           = Query(default_factory=lambda: datetime.today().year),
    charge_point_id: Optional[str] = Query(None),
    db: Session         = Depends(get_db),
    _: User             = Depends(get_current_user),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="openpyxl tidak terinstall")

    wb = openpyxl.Workbook()
    ws = wb.active

    # Styling
    HEADER_FILL  = PatternFill("solid", fgColor="1A1F2E")
    HEADER_FONT  = Font(bold=True, color="10B981", size=10)
    ALT_FILL     = PatternFill("solid", fgColor="0F1320")
    THIN_BORDER  = Border(
        bottom=Side(style="thin", color="2D3748"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    def style_header_row(row_cells):
        for cell in row_cells:
            cell.fill       = HEADER_FILL
            cell.font       = HEADER_FONT
            cell.alignment  = center_align
            cell.border     = THIN_BORDER

    def style_data_row(row_cells, alt: bool):
        for cell in row_cells:
            if alt:
                cell.fill = ALT_FILL
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    if report_type == "daily":
        ws.title = "Laporan Harian"
        headers  = ["Tanggal", "Charge Point ID", "Nama CP",
                    "Jumlah Transaksi", "Energi (kWh)", "Pendapatan (Rp)"]
        ws.append(headers)
        style_header_row(ws[1])

        rows = report_daily(date_from, date_to, charge_point_id, db, _)
        for i, r in enumerate(rows):
            ws.append([
                r.report_date, r.charge_point_id, r.charge_point_name,
                r.total_transactions,
                round(r.total_energy_kwh, 3),
                round(r.total_revenue, 0),
            ])
            style_data_row(list(ws.iter_rows(min_row=i+2, max_row=i+2))[0], i % 2 == 1)

        col_widths = [14, 20, 28, 18, 16, 20]
        filename   = f"laporan_harian_{date_from}_{date_to}.xlsx"

    else:
        ws.title = "Laporan Bulanan"
        headers  = ["Tahun", "Bulan", "Charge Point ID", "Nama CP",
                    "Jumlah Transaksi", "Energi (kWh)", "Rata-rata kWh/Tx", "Pendapatan (Rp)"]
        ws.append(headers)
        style_header_row(ws[1])

        rows = report_monthly(year, charge_point_id, db, _)
        for i, r in enumerate(rows):
            ws.append([
                r.year, r.month_name, r.charge_point_id, r.charge_point_name,
                r.total_transactions,
                round(r.total_energy_kwh, 3),
                round(r.avg_energy_kwh, 3),
                round(r.total_revenue, 0),
            ])
            style_data_row(list(ws.iter_rows(min_row=i+2, max_row=i+2))[0], i % 2 == 1)

        col_widths = [8, 14, 20, 28, 18, 16, 18, 20]
        filename   = f"laporan_bulanan_{year}.xlsx"

    # Set lebar kolom
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
